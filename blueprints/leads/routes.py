import csv
import io
import json
import os
import re
import tempfile
import uuid
from datetime import datetime
from flask import render_template, request, redirect, url_for, flash, jsonify, session
from flask_login import login_required, current_user
from . import leads_bp
from extensions import db
from models import Lead, Campaign, EnrolledLead, EnrolledStatus, Template, SendLog, SendStatus


@leads_bp.route('/')
@login_required
def pool():
    q = request.args.get('q', '').strip()
    status_filter = request.args.get('status', '')
    company_filter = request.args.get('company', '').strip()

    query = Lead.query.filter_by(user_id=current_user.id)

    if q:
        query = query.filter(
            (Lead.first_name.ilike(f'%{q}%')) |
            (Lead.last_name.ilike(f'%{q}%')) |
            (Lead.email.ilike(f'%{q}%')) |
            (Lead.company.ilike(f'%{q}%'))
        )

    if company_filter:
        query = query.filter(Lead.company.ilike(f'%{company_filter}%'))

    leads = query.order_by(Lead.created_at.desc()).all()

    # Campaigns and email accounts for bulk enroll
    from models import EmailAccount
    campaigns = Campaign.query.filter_by(user_id=current_user.id).order_by(Campaign.name).all()
    email_accounts = EmailAccount.query.filter_by(user_id=current_user.id, active=True).order_by(EmailAccount.email_address).all()

    return render_template('leads/pool.html', leads=leads, campaigns=campaigns,
                           email_accounts=email_accounts, q=q, status_filter=status_filter)


def _upload_tmp_path(user_id, upload_id):
    safe = ''.join(c for c in upload_id if c.isalnum() or c == '-')
    return os.path.join(tempfile.gettempdir(), f'leadflow_{user_id}_{safe}.json')


@leads_bp.route('/upload', methods=['GET', 'POST'])
@login_required
def upload():
    if request.method == 'POST':
        upload_id = request.form.get('upload_id', '').strip()

        if upload_id:
            # ── Step 2: user submitted mapping form ──
            tmp_path = _upload_tmp_path(current_user.id, upload_id)
            try:
                with open(tmp_path, 'r', encoding='utf-8') as f:
                    rows = json.load(f)
                os.unlink(tmp_path)
            except (FileNotFoundError, OSError, ValueError):
                flash('Upload session expired. Please upload again.', 'danger')
                return redirect(url_for('leads.upload'))

            mapping = {
                'email':        request.form.get('col_email', ''),
                'first_name':   request.form.get('col_first_name', ''),
                'last_name':    request.form.get('col_last_name', ''),
                'company':      request.form.get('col_company', ''),
                'title':        request.form.get('col_title', ''),
                'website':      request.form.get('col_website', ''),
                'phone':        request.form.get('col_phone', ''),
                'linkedin_url': request.form.get('col_linkedin', ''),
                'signal_1':     request.form.get('col_signal_1', ''),
                'signal_2':     request.form.get('col_signal_2', ''),
            }

            if not mapping['email']:
                flash('Email column is required.', 'danger')
                headers = [k for k in (rows[0].keys() if rows else []) if k]
                return render_template('leads/upload.html', rows=rows[:5], headers=headers,
                                       mapping=mapping, upload_id=upload_id,
                                       total_rows=len(rows))

            # Custom column pairs submitted as parallel lists
            custom_names = request.form.getlist('custom_col_name')
            custom_csvs  = request.form.getlist('custom_col_csv')
            custom_mappings = [
                (n.strip(), c)
                for n, c in zip(custom_names, custom_csvs)
                if n.strip() and c
            ]

            imported = duplicates = invalid = 0
            new_leads = []
            for row in rows:
                email = row.get(mapping['email'], '').strip()
                if not email or '@' not in email:
                    invalid += 1
                    continue

                existing = Lead.query.filter_by(
                    user_id=current_user.id, email=email
                ).first()
                if existing:
                    duplicates += 1
                    continue

                def gv(col_key):
                    col = mapping.get(col_key, '')
                    return row.get(col, '').strip() if col else ''

                extra = {
                    name: row.get(csv_col, '').strip()
                    for name, csv_col in custom_mappings
                    if row.get(csv_col, '').strip()
                }

                lead = Lead(
                    user_id=current_user.id,
                    email=email,
                    first_name=gv('first_name'),
                    last_name=gv('last_name'),
                    company=gv('company'),
                    title=gv('title'),
                    website=gv('website'),
                    phone=gv('phone'),
                    linkedin_url=gv('linkedin_url'),
                    signal_1=gv('signal_1'),
                    signal_2=gv('signal_2'),
                    extra_data=json.dumps(extra) if extra else None,
                    source='upload',
                )
                db.session.add(lead)
                new_leads.append(lead)
                imported += 1

            # Optional campaign enrollment
            campaign_id = request.form.get('campaign_id', '').strip()
            new_campaign_name = request.form.get('new_campaign_name', '').strip()
            campaign = None

            if campaign_id == '__new__' and new_campaign_name:
                campaign = Campaign(user_id=current_user.id, name=new_campaign_name)
                db.session.add(campaign)
            elif campaign_id and campaign_id != '__new__':
                campaign = Campaign.query.filter_by(
                    id=int(campaign_id), user_id=current_user.id
                ).first()

            # Flush to get IDs before creating EnrolledLead rows
            db.session.flush()

            enrolled_count = 0
            if campaign:
                for lead in new_leads:
                    el = EnrolledLead(
                        campaign_id=campaign.id,
                        lead_id=lead.id,
                        status=EnrolledStatus.ACTIVE,
                    )
                    db.session.add(el)
                    enrolled_count += 1

            db.session.commit()

            parts = [f'Imported {imported} leads.']
            if campaign and enrolled_count:
                parts.append(f'Enrolled {enrolled_count} in "{campaign.name}".')
            if duplicates:
                parts.append(f'{duplicates} already in your pool (skipped).')
            if invalid:
                parts.append(f'{invalid} had missing or invalid emails (skipped).')
            flash(' '.join(parts), 'success')
            return redirect(url_for('leads.pool'))

        else:
            # ── Step 1: parse file, save to temp, show mapping form ──
            file = request.files.get('file')
            if not file or not file.filename:
                flash('Please select a file.', 'danger')
                return redirect(url_for('leads.upload'))

            filename = file.filename.lower()
            rows = []

            try:
                if filename.endswith('.csv'):
                    content = file.stream.read().decode('utf-8-sig', errors='replace')
                    reader = csv.DictReader(io.StringIO(content))
                    rows = [dict(r) for r in reader]
                elif filename.endswith(('.xlsx', '.xls')):
                    import openpyxl
                    wb = openpyxl.load_workbook(file.stream)
                    ws = wb.active
                    hdrs = [str(c.value or '').strip() for c in next(ws.iter_rows(max_row=1))]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        rows.append(dict(zip(hdrs, [str(v or '') for v in row])))
                else:
                    flash('Only CSV and Excel files are supported.', 'danger')
                    return redirect(url_for('leads.upload'))
            except Exception as e:
                flash(f'Error reading file: {e}', 'danger')
                return redirect(url_for('leads.upload'))

            if not rows:
                flash('File is empty or could not be read.', 'warning')
                return redirect(url_for('leads.upload'))

            upload_id = str(uuid.uuid4())
            tmp_path = _upload_tmp_path(current_user.id, upload_id)
            with open(tmp_path, 'w', encoding='utf-8') as f:
                json.dump(rows, f, default=str)

            headers = [k for k in rows[0].keys() if k]

            def _auto(kw):
                return next((h for h in headers if kw in h.lower()), '')

            auto_map = {
                'col_email':      _auto('email'),
                'col_first_name': _auto('first'),
                'col_last_name':  _auto('last'),
                'col_company':    _auto('company'),
                'col_title':      _auto('title'),
                'col_website':    _auto('website'),
                'col_phone':      _auto('phone'),
                'col_linkedin':   _auto('linkedin'),
                'col_signal_1':   next((h for h in headers if 'signal' in h.lower() and '1' in h), ''),
                'col_signal_2':   next((h for h in headers if 'signal' in h.lower() and '2' in h), ''),
            }

            campaigns = Campaign.query.filter_by(
                user_id=current_user.id
            ).order_by(Campaign.name).all()

            return render_template('leads/upload.html',
                                   rows=rows[:5], headers=headers,
                                   mapping=auto_map, upload_id=upload_id,
                                   total_rows=len(rows),
                                   campaigns=campaigns)

    return render_template('leads/upload.html', rows=[], headers=[], mapping={})


@leads_bp.route('/new', methods=['GET', 'POST'])
@login_required
def new_lead():
    if request.method == 'POST':
        email = request.form.get('email', '').strip()
        if not email:
            flash('Email is required.', 'danger')
            return render_template('leads/new.html')

        existing = Lead.query.filter_by(user_id=current_user.id, email=email).first()
        if existing:
            flash('A lead with that email already exists.', 'warning')
            return redirect(url_for('leads.detail', lead_id=existing.id))

        lead = Lead(
            user_id=current_user.id,
            email=email,
            first_name=request.form.get('first_name', '').strip(),
            last_name=request.form.get('last_name', '').strip(),
            company=request.form.get('company', '').strip(),
            title=request.form.get('title', '').strip(),
            website=request.form.get('website', '').strip(),
            phone=request.form.get('phone', '').strip(),
            linkedin_url=request.form.get('linkedin_url', '').strip(),
            signal_1=request.form.get('signal_1', '').strip(),
            signal_2=request.form.get('signal_2', '').strip(),
            signal_3=request.form.get('signal_3', '').strip(),
            account_grade=request.form.get('account_grade', 'B'),
            notes=request.form.get('notes', '').strip(),
            source='manual',
        )
        db.session.add(lead)
        db.session.commit()
        flash(f'Lead {lead.full_name} added!', 'success')
        return redirect(url_for('leads.detail', lead_id=lead.id))

    return render_template('leads/new.html')


@leads_bp.route('/<int:lead_id>')
@login_required
def detail(lead_id):
    lead = Lead.query.filter_by(id=lead_id, user_id=current_user.id).first_or_404()
    enrolled = EnrolledLead.query.filter_by(lead_id=lead_id).all()
    campaigns = Campaign.query.filter_by(user_id=current_user.id).all()
    return render_template('leads/detail.html', lead=lead, enrolled=enrolled, campaigns=campaigns)


@leads_bp.route('/<int:lead_id>', methods=['PATCH', 'POST'])
@login_required
def update_lead(lead_id):
    lead = Lead.query.filter_by(id=lead_id, user_id=current_user.id).first_or_404()

    data = request.get_json(silent=True) or request.form

    for field in ['first_name', 'last_name', 'email', 'company', 'title',
                  'website', 'phone', 'linkedin_url', 'signal_1', 'signal_2',
                  'signal_3', 'account_grade', 'notes']:
        if field in data:
            setattr(lead, field, data[field])

    if 'do_not_contact' in data:
        val = data['do_not_contact']
        lead.do_not_contact = val in (True, 'true', '1', 'on')

    db.session.commit()

    if request.is_json:
        return jsonify({'ok': True})
    flash('Lead updated.', 'success')
    return redirect(url_for('leads.detail', lead_id=lead_id))


@leads_bp.route('/<int:lead_id>/delete', methods=['POST'])
@login_required
def delete_lead(lead_id):
    lead = Lead.query.filter_by(id=lead_id, user_id=current_user.id).first_or_404()
    db.session.delete(lead)
    db.session.commit()
    flash('Lead deleted.', 'success')
    return redirect(url_for('leads.pool'))


@leads_bp.route('/<int:lead_id>/enroll', methods=['POST'])
@login_required
def enroll_lead(lead_id):
    lead = Lead.query.filter_by(id=lead_id, user_id=current_user.id).first_or_404()
    campaign_id = request.form.get('campaign_id') or (request.get_json(silent=True) or {}).get('campaign_id')

    if not campaign_id:
        flash('Please select a campaign.', 'warning')
        return redirect(url_for('leads.detail', lead_id=lead_id))

    campaign = Campaign.query.filter_by(id=int(campaign_id), user_id=current_user.id).first_or_404()

    existing = EnrolledLead.query.filter_by(campaign_id=campaign.id, lead_id=lead.id).first()
    if existing:
        flash('Lead is already enrolled in that campaign.', 'warning')
        return redirect(url_for('leads.detail', lead_id=lead_id))

    el = EnrolledLead(
        campaign_id=campaign.id,
        lead_id=lead.id,
        status=EnrolledStatus.ACTIVE,
    )
    db.session.add(el)
    db.session.commit()
    flash(f'Lead enrolled in campaign "{campaign.name}".', 'success')
    return redirect(url_for('leads.detail', lead_id=lead_id))


@leads_bp.route('/ai-prompt', methods=['GET', 'POST'])
@login_required
def ai_prompt():
    """Show AI prompt builder page. Accepts lead IDs from pool (POST) or campaign (GET)."""
    if request.method == 'POST':
        lead_ids = request.form.getlist('lead_ids')
        campaign_id = request.form.get('campaign_id', '')
    else:
        lead_ids = request.args.getlist('lead_ids')
        campaign_id = request.args.get('campaign_id', '')

    # If campaign_id given and no lead_ids, load all enrolled leads from that campaign
    if campaign_id and not lead_ids:
        campaign = Campaign.query.filter_by(id=int(campaign_id), user_id=current_user.id).first_or_404()
        enrolled = EnrolledLead.query.filter_by(campaign_id=campaign.id).all()
        lead_ids = [str(el.lead_id) for el in enrolled]

    if not lead_ids:
        flash('Select at least one lead first.', 'warning')
        return redirect(url_for('leads.pool'))

    leads = Lead.query.filter(
        Lead.id.in_([int(x) for x in lead_ids]),
        Lead.user_id == current_user.id
    ).all()

    if not leads:
        flash('No valid leads found.', 'warning')
        return redirect(url_for('leads.pool'))

    campaigns = Campaign.query.filter_by(user_id=current_user.id).order_by(Campaign.name).all()
    templates = Template.query.filter(
        (Template.user_id == current_user.id) | (Template.is_builtin == True)
    ).order_by(Template.touch_type, Template.name).all()

    # Build per-campaign sequence step data for the JS
    _touch_labels = {
        'opener': 'Opener', 'observation': 'Opening Touch',
        'hypothesis': 'Follow-up 2', 'proof': 'Follow-up 3 — Proof',
        'soft_close': 'Soft Close', 'breakup': 'Breakup',
        'not_now': 'Not Now Reply', 'link_clicked': 'Link Clicked',
        'reply_received': 'Reply Received', 'out_of_office': 'Out of Office',
    }
    campaign_seq_data = {}
    for c in campaigns:
        if c.sequence_id and c.sequence:
            seq_steps = []
            for step in sorted(c.sequence.steps.all(), key=lambda s: s.day_offset):
                if step.channel != 'Email':
                    continue
                st = step.step_templates.filter_by(is_active=True).first()
                pinned = st.template if st else None
                touch = step.template_slot
                seq_steps.append({
                    'touch_type': touch,
                    'day': step.day_offset,
                    'template_id': pinned.id if pinned else None,
                    'template_name': pinned.name if pinned else None,
                    'label': _touch_labels.get(touch, touch.replace('_', ' ').title()),
                    'desc': f'Day {step.day_offset}' + (f' — {pinned.name}' if pinned else ''),
                })
            if seq_steps:
                campaign_seq_data[c.id] = seq_steps

    return render_template('leads/ai_prompt.html',
                           leads=leads,
                           lead_ids=[str(l.id) for l in leads],
                           campaign_id=campaign_id,
                           campaigns=campaigns,
                           templates=templates,
                           campaign_seq_data=campaign_seq_data)


@leads_bp.route('/ai-prompt/generate', methods=['POST'])
@login_required
def ai_prompt_generate():
    """Return the prompt text as JSON given lead IDs + touch type + template."""
    lead_ids = request.form.getlist('lead_ids')
    touch_type = request.form.get('touch_type', 'observation')
    template_id = request.form.get('template_id', '')

    leads = Lead.query.filter(
        Lead.id.in_([int(x) for x in lead_ids]),
        Lead.user_id == current_user.id
    ).all()

    if not leads:
        return jsonify({'error': 'No leads found'}), 400

    # Template guidance
    template_subject = ''
    template_body = ''
    if template_id:
        t = Template.query.get(int(template_id))
        if t and (t.is_builtin or t.user_id == current_user.id):
            template_subject = t.subject or ''
            template_body = t.body or ''

    touch_labels = {
        'observation': 'Opening Touch — Signal Observation',
        'hypothesis': 'Second Touch — Hypothesis',
        'proof': 'Third Touch — Proof / Case Study',
        'soft_close': 'Soft Close',
        'breakup': 'Breakup Email',
        'not_now': 'Not Now Follow-up',
        'link_clicked': 'Link Clicked Follow-up',
        'reply_received': 'Reply Received Response',
    }
    touch_label = touch_labels.get(touch_type, touch_type.replace('_', ' ').title())

    # Build lead blocks
    lead_blocks = []
    for lead in leads:
        parts = [f'Email: {lead.email}']
        if lead.first_name:  parts.append(f'First Name: {lead.first_name}')
        if lead.last_name:   parts.append(f'Last Name: {lead.last_name}')
        if lead.company:     parts.append(f'Company: {lead.company}')
        if lead.title:       parts.append(f'Title: {lead.title}')
        if lead.website:     parts.append(f'Website: {lead.website}')
        if lead.signal_1:    parts.append(f'Signal 1: {lead.signal_1}')
        if lead.signal_2:    parts.append(f'Signal 2: {lead.signal_2}')
        lead_blocks.append('\n'.join(parts))

    template_section = ''
    if template_body:
        template_section = (
            f'\nTEMPLATE TO ADAPT (use the style and structure — do not copy verbatim):\n'
            f'Subject template: {template_subject}\n'
            f'Body template:\n{template_body}\n'
        )

    prompt = f"""You are writing personalized cold outreach emails for a B2B sales professional.

Write a "{touch_label}" email for EVERY lead listed below.

RULES:
- Under 100 words per email
- Plain text only — no bullet points, no HTML, no markdown formatting
- Conversational and human — not salesy or corporate
- Personalize using the lead's name, company, title, and any signals provided
- Replace any {{{{placeholders}}}} with actual lead information
- Do not invent facts you don't have
{template_section}
OUTPUT FORMAT — copy this pattern exactly for every lead, no exceptions:

---LEAD:{{email address}}---
SUBJECT: {{subject line}}
BODY:
{{email body}}
---END---

Write all {len(leads)} emails now. Start immediately with the first ---LEAD:--- block.

LEADS ({len(leads)} total):

""" + '\n\n'.join(f'[Lead {i + 1}]\n{block}' for i, block in enumerate(lead_blocks))

    return jsonify({'prompt': prompt, 'count': len(leads)})


@leads_bp.route('/ai-prompt/import', methods=['POST'])
@login_required
def ai_prompt_import():
    """Parse the AI output and save drafts or send immediately."""
    from email_service import send_email

    raw_output = request.form.get('ai_output', '').strip()
    campaign_id = request.form.get('campaign_id', '').strip()
    touch_type = request.form.get('touch_type', 'observation')
    action = request.form.get('action', 'draft')  # 'draft' or 'send'
    lead_ids_raw = request.form.get('lead_ids_json', '[]')

    if not raw_output:
        flash('Paste the AI output before importing.', 'warning')
        return redirect(url_for('leads.pool'))

    if not campaign_id:
        flash('A campaign must be selected to import emails.', 'warning')
        return redirect(url_for('leads.pool'))

    campaign = Campaign.query.filter_by(id=int(campaign_id), user_id=current_user.id).first_or_404()

    # Parse structured output: ---LEAD:email--- ... SUBJECT: ... BODY: ... ---END---
    pattern = r'---LEAD:\s*(.*?)\s*---\s*SUBJECT:\s*(.*?)\s*BODY:\s*(.*?)\s*---END---'
    matches = re.findall(pattern, raw_output, re.DOTALL | re.IGNORECASE)

    if not matches:
        flash(
            'Could not parse the AI output. Make sure you copied the full response '
            'including all ---LEAD:--- and ---END--- markers.',
            'danger'
        )
        return redirect(url_for('campaigns.detail', campaign_id=campaign.id, tab='review'))

    # Find a template for this touch type (for logging)
    template = Template.query.filter(
        (Template.user_id == current_user.id) | (Template.is_builtin == True),
        Template.touch_type == touch_type
    ).first()

    # Find the matching sequence step
    step = None
    if campaign.sequence:
        for s in campaign.sequence.steps.all():
            if s.template_slot == touch_type and s.channel == 'Email':
                step = s
                break

    saved = 0
    sent_ok = 0
    skipped = 0
    errors = []

    for email_addr, subject, body in matches:
        email_addr = email_addr.strip().lower()
        subject = subject.strip()
        body = body.strip()

        if not email_addr or not subject or not body:
            skipped += 1
            continue

        lead = Lead.query.filter_by(user_id=current_user.id, email=email_addr).first()
        if not lead:
            skipped += 1
            errors.append(f'{email_addr}: lead not found in your account')
            continue

        el = EnrolledLead.query.filter_by(campaign_id=campaign.id, lead_id=lead.id).first()
        if not el:
            skipped += 1
            errors.append(f'{email_addr}: not enrolled in this campaign')
            continue

        if action == 'send':
            # Send immediately via an active account
            accounts = __import__('models', fromlist=['EmailAccount']).EmailAccount.query.filter_by(
                user_id=current_user.id, active=True
            ).all()
            account = accounts[sent_ok % len(accounts)] if accounts else None

            if not account:
                skipped += 1
                errors.append(f'{email_addr}: no active email account')
                continue

            success, error = send_email(account, lead.email, subject, body, current_user)
            log = SendLog(
                enrolled_lead_id=el.id,
                step_id=step.id if step else None,
                template_id=template.id if template else None,
                variant_label='AI-Prompt',
                subject=subject,
                body_snippet=body[:2000],
                status=SendStatus.SENT if success else SendStatus.FAILED,
                sent_at=datetime.utcnow() if success else None,
                from_account_id=account.id if success else None,
            )
            db.session.add(log)
            if success:
                sent_ok += 1
            else:
                skipped += 1
                errors.append(f'{email_addr}: send failed — {error}')
        else:
            # Save as draft for review
            log = SendLog(
                enrolled_lead_id=el.id,
                step_id=step.id if step else None,
                template_id=template.id if template else None,
                variant_label='AI-Prompt',
                subject=subject,
                body_snippet=body[:2000],
                status=SendStatus.DRAFT,
            )
            db.session.add(log)
            saved += 1

    db.session.commit()

    if action == 'send':
        if sent_ok:
            flash(f'Sent {sent_ok} email{"s" if sent_ok != 1 else ""}. {skipped} skipped.', 'success')
        else:
            flash('No emails could be sent. Check the errors below.', 'danger')
    else:
        if saved:
            flash(
                f'{saved} draft{"s" if saved != 1 else ""} saved — review them in the Review Queue tab. '
                f'{skipped} skipped.',
                'success'
            )
        else:
            flash('No drafts could be saved. Make sure the leads are enrolled in this campaign.', 'danger')

    if errors and skipped:
        flash('Skipped: ' + '; '.join(errors[:5]) + (f' (+{len(errors)-5} more)' if len(errors) > 5 else ''), 'warning')

    return redirect(url_for('campaigns.detail', campaign_id=campaign.id, tab='review'))


@leads_bp.route('/bulk-delete', methods=['POST'])
@login_required
def bulk_delete():
    lead_ids = request.form.getlist('lead_ids')
    if not lead_ids:
        flash('No leads selected.', 'warning')
        return redirect(url_for('leads.pool'))

    deleted = 0
    for lid in lead_ids:
        lead = Lead.query.filter_by(id=int(lid), user_id=current_user.id).first()
        if lead:
            db.session.delete(lead)
            deleted += 1

    db.session.commit()
    flash(f'Deleted {deleted} lead{"s" if deleted != 1 else ""}.', 'success')
    return redirect(url_for('leads.pool'))


@leads_bp.route('/bulk-enroll', methods=['POST'])
@login_required
def bulk_enroll():
    lead_ids = request.form.getlist('lead_ids')
    campaign_id = request.form.get('campaign_id', '').strip()
    new_campaign_name = request.form.get('new_campaign_name', '').strip()

    if not lead_ids:
        flash('Select at least one lead.', 'warning')
        return redirect(url_for('leads.pool'))

    if not campaign_id:
        flash('Select a campaign or create a new one.', 'warning')
        return redirect(url_for('leads.pool'))

    if campaign_id == '__new__':
        if not new_campaign_name:
            flash('Enter a name for the new campaign.', 'warning')
            return redirect(url_for('leads.pool'))
        campaign = Campaign(user_id=current_user.id, name=new_campaign_name)
        db.session.add(campaign)
        db.session.flush()
    else:
        campaign = Campaign.query.filter_by(id=int(campaign_id), user_id=current_user.id).first_or_404()

    # start_step: how many touches have already been sent outside LeadFlow
    try:
        start_step = max(0, int(request.form.get('start_step', 0)))
    except (ValueError, TypeError):
        start_step = 0

    # Pinned sending account
    from_account_id_raw = request.form.get('from_account_id', '').strip()
    from_account_id = int(from_account_id_raw) if from_account_id_raw.isdigit() else None

    enrolled = 0
    skipped = 0
    for lid in lead_ids:
        lead = Lead.query.filter_by(id=int(lid), user_id=current_user.id).first()
        if not lead:
            continue
        existing = EnrolledLead.query.filter_by(campaign_id=campaign.id, lead_id=lead.id).first()
        if existing:
            skipped += 1
            continue
        el = EnrolledLead(
            campaign_id=campaign.id,
            lead_id=lead.id,
            status=EnrolledStatus.ACTIVE,
            current_step=start_step,
            from_account_id=from_account_id,
        )
        db.session.add(el)
        enrolled += 1

    db.session.commit()
    step_msg = f' (starting at touch {start_step + 1})' if start_step > 0 else ''
    flash(f'Enrolled {enrolled} leads in "{campaign.name}"{step_msg}. {skipped} already enrolled.', 'success')
    return redirect(url_for('leads.pool'))


@leads_bp.route('/pre-load', methods=['GET', 'POST'])
@login_required
def pre_load():
    """Batch pre-load email drafts for non-API-key users."""
    from models import EmailAccount, Sequence, SequenceStep, StepTemplate, Template
    campaigns = Campaign.query.filter_by(user_id=current_user.id).order_by(Campaign.name).all()

    if request.method == 'GET':
        campaign_id = request.args.get('campaign_id', '')
        try:
            limit = max(1, min(500, int(request.args.get('limit', 50))))
        except (ValueError, TypeError):
            limit = 50

        slots = []
        if campaign_id:
            campaign = Campaign.query.filter_by(id=int(campaign_id), user_id=current_user.id).first_or_404()
            if campaign.sequence_id:
                from datetime import date, timedelta
                from scheduler_jobs import _step_jitter
                sequence = campaign.sequence
                steps = sorted(sequence.steps.all(), key=lambda s: s.day_offset)

                active_leads = EnrolledLead.query.filter_by(
                    campaign_id=campaign.id, status=EnrolledStatus.ACTIVE
                ).all()

                today = date.today()
                for el in active_leads:
                    if len(slots) >= limit:
                        break
                    lead = el.lead
                    for step in steps:
                        jitter = _step_jitter(el.id, step.id)
                        due_date = el.enrolled_at.date() + timedelta(days=step.day_offset + jitter)
                        if today < due_date:
                            continue
                        if step.channel != 'Email':
                            continue
                        existing = SendLog.query.filter_by(
                            enrolled_lead_id=el.id, step_id=step.id
                        ).first()
                        if existing:
                            continue
                        # Build prompt
                        step_templates = [st for st in step.step_templates.all() if st.is_active]
                        if step_templates:
                            tmpl = step_templates[0].template
                        else:
                            # Prefer user's own template for this touch type, then builtin
                            from models import Campaign as _Camp
                            _camp_user = _Camp.query.get(c.id).user_id
                            tmpl = (
                                Template.query.filter_by(user_id=_camp_user, touch_type=step.template_slot).first()
                                or Template.query.filter_by(touch_type=step.template_slot, is_builtin=True).first()
                            )
                        if not tmpl:
                            continue
                        slots.append({
                            'enrolled_lead_id': el.id,
                            'step_id': step.id,
                            'template_id': tmpl.id if tmpl else None,
                            'lead_name': f'{lead.first_name or ""} {lead.last_name or ""}'.strip() or lead.email,
                            'company': lead.company or '',
                            'email': lead.email,
                            'touch': step.template_slot.replace('_', ' ').title(),
                            'template_subject': (tmpl.subject or '').replace('{{first_name}}', lead.first_name or '').replace('{{company}}', lead.company or ''),
                            'template_body': (tmpl.body or '').replace('{{first_name}}', lead.first_name or '').replace('{{company}}', lead.company or ''),
                        })
                        break  # one pending step per lead

        return render_template('leads/pre_load.html',
                               campaigns=campaigns,
                               selected_campaign_id=campaign_id,
                               limit=limit,
                               slots=slots)

    # POST — save the filled drafts
    enrolled_lead_ids = request.form.getlist('enrolled_lead_id')
    step_ids = request.form.getlist('step_id')
    template_ids = request.form.getlist('template_id')
    subjects = request.form.getlist('subject')
    bodies = request.form.getlist('body')

    saved = 0
    for i, el_id in enumerate(enrolled_lead_ids):
        body_text = bodies[i].strip() if i < len(bodies) else ''
        subject_text = subjects[i].strip() if i < len(subjects) else ''
        if not body_text:
            continue  # skip blanks
        el = EnrolledLead.query.filter_by(id=int(el_id), status=EnrolledStatus.ACTIVE).first()
        if not el or el.lead.user_id != current_user.id:
            continue
        step_id = int(step_ids[i]) if i < len(step_ids) else None
        tmpl_id = int(template_ids[i]) if i < len(template_ids) and template_ids[i] else None
        existing = SendLog.query.filter_by(enrolled_lead_id=el.id, step_id=step_id).first()
        if existing:
            continue
        draft = SendLog(
            enrolled_lead_id=el.id,
            step_id=step_id,
            template_id=tmpl_id,
            variant_label='A',
            subject=subject_text,
            body_snippet=body_text[:4000],
            status='queued',
        )
        db.session.add(draft)
        saved += 1

    db.session.commit()
    flash(f'{saved} email drafts queued — the scheduler will send them at your daily rate automatically.', 'success')
    return redirect(url_for('leads.pool'))
